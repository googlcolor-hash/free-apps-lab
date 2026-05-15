from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / 'data'
OUT_DIR = ROOT / 'deliverables'
APPS_JSON = DATA_DIR / 'apps.json'
STATS_JSON = DATA_DIR / 'stats.json'
OUT_XLSX = OUT_DIR / 'free_apps_catalog.xlsx'

REQUIRED_FIELDS = [
    'id',
    'name',
    'url',
    'description',
    'section_l1',
    'section_l2',
    'platforms',
    'is_open_source',
    'is_recommended',
    'source_file',
    'source_line',
]


HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _load_apps() -> list[dict]:
    if not APPS_JSON.exists():
        raise SystemExit(f'WAIT: missing {APPS_JSON}')
    data = json.loads(APPS_JSON.read_text(encoding='utf-8'))
    if not isinstance(data, list) or not data:
        raise SystemExit('WAIT: apps.json is empty or invalid')
    return data


def _normalize_platforms(raw: object) -> list[str]:
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    if raw is None:
        return []
    val = str(raw).strip()
    if not val:
        return []
    if ',' in val:
        return [p.strip() for p in val.split(',') if p.strip()]
    return [val]


def _derive_stats(apps: list[dict]) -> dict:
    categories = Counter()
    platforms = Counter()
    flags = {
        'total': len(apps),
        'open_source_true': 0,
        'recommended_true': 0,
    }

    for app in apps:
        cat = str(app.get('section_l1') or 'Uncategorized')
        categories[cat] += 1

        for platform in _normalize_platforms(app.get('platforms')):
            platforms[platform] += 1

        if bool(app.get('is_open_source')):
            flags['open_source_true'] += 1
        if bool(app.get('is_recommended')):
            flags['recommended_true'] += 1

    top_categories = categories.most_common(10)

    return {
        'categories': dict(categories),
        'platforms': dict(platforms),
        'flags': flags,
        'top_categories': [
            {'category': name, 'count': count} for name, count in top_categories
        ],
    }


def _load_or_build_stats(apps: list[dict]) -> dict:
    if STATS_JSON.exists():
        try:
            return json.loads(STATS_JSON.read_text(encoding='utf-8'))
        except json.JSONDecodeError:
            pass

    stats = _derive_stats(apps)
    STATS_JSON.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding='utf-8')
    return stats


def _style_header(ws, row_idx: int = 1):
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _autosize(ws, max_width: int = 64):
    for column_cells in ws.columns:
        values = [str(c.value) for c in column_cells if c.value is not None]
        width = min(max((len(v) for v in values), default=10) + 2, max_width)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def build_workbook(apps: list[dict], stats: dict) -> Workbook:
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = 'Raw'

    ws_raw.append(REQUIRED_FIELDS)
    for app in apps:
        row = []
        for f in REQUIRED_FIELDS:
            value = app.get(f)
            if f == 'platforms':
                value = ', '.join(_normalize_platforms(value))
            row.append(value)
        ws_raw.append(row)
    _style_header(ws_raw)
    _autosize(ws_raw)

    ws_categories = wb.create_sheet('Categories')
    ws_categories.append(['Категория', 'Количество'])
    for k, v in sorted(stats.get('categories', {}).items(), key=lambda x: x[0].lower()):
        ws_categories.append([k, v])
    _style_header(ws_categories)
    _autosize(ws_categories)

    ws_platforms = wb.create_sheet('Platforms')
    ws_platforms.append(['Платформа', 'Количество'])
    for k, v in sorted(stats.get('platforms', {}).items(), key=lambda x: x[0].lower()):
        ws_platforms.append([k, v])
    _style_header(ws_platforms)
    _autosize(ws_platforms)

    flags = stats.get('flags', {})
    ws_flags = wb.create_sheet('Flags')
    ws_flags.append(['Метрика', 'Значение'])
    ws_flags.append(['Всего приложений', flags.get('total', len(apps))])
    ws_flags.append(['Open Source = true', flags.get('open_source_true', 0)])
    ws_flags.append(['Recommended = true', flags.get('recommended_true', 0)])
    _style_header(ws_flags)
    _autosize(ws_flags)

    ws_top = wb.create_sheet('TopCategories')
    ws_top.append(['Категория', 'Количество'])
    for item in stats.get('top_categories', []):
        ws_top.append([item.get('category'), item.get('count')])
    _style_header(ws_top)
    _autosize(ws_top)

    return wb


def main():
    apps = _load_apps()
    stats = _load_or_build_stats(apps)
    wb = build_workbook(apps, stats)
    wb.save(OUT_XLSX)
    print(f'OK: created {OUT_XLSX}')


if __name__ == '__main__':
    main()
